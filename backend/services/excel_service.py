import io
import re
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Any, Tuple

from openpyxl import Workbook, load_workbook

TEMPLATES = {
    "product": {
        "headers": ["产品名称", "规格型号", "默认单价", "备注"],
        "required": ["产品名称"],
    },
    "customer": {
        "headers": ["客户名称", "联系人", "联系电话", "地址", "备注"],
        "required": ["客户名称", "联系电话"],
    },
}


def generate_template(entity_type: str) -> bytes:
    tpl = TEMPLATES.get(entity_type)
    if not tpl:
        raise ValueError(f"不支持的实体类型: {entity_type}")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{entity_type}_导入模板"
    ws.append(tpl["headers"])

    if entity_type == "product":
        ws.append(["示例产品A", "X-100", "99.50", "备注示例"])
    elif entity_type == "customer":
        ws.append(["示例客户公司", "联系人姓名", "13800138000", "地址示例", "备注示例"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _validate_product_row(row: Dict[str, str], row_num: int) -> List[str]:
    """校验单行产品数据，返回错误列表"""
    errors = []

    name = row.get("产品名称", "").strip()
    if not name:
        errors.append(f"第{row_num}行: 产品名称不能为空")
    elif len(name) > 200:
        errors.append(f"第{row_num}行: 产品名称不能超过200个字符")

    model = row.get("规格型号", "").strip()
    if len(model) > 200:
        errors.append(f"第{row_num}行: 规格型号不能超过200个字符")

    price_str = row.get("默认单价", "").strip()
    if price_str:
        try:
            price = Decimal(price_str)
            if price < 0:
                errors.append(f"第{row_num}行: 默认单价不能为负数")
        except InvalidOperation:
            errors.append(f"第{row_num}行: 默认单价格式无效，请输入数字（如 99.50）")

    return errors


def _validate_customer_row(row: Dict[str, str], row_num: int) -> List[str]:
    """校验单行客户数据，返回错误列表"""
    errors = []

    name = row.get("客户名称", "").strip()
    if not name:
        errors.append(f"第{row_num}行: 客户名称不能为空")
    elif len(name) > 200:
        errors.append(f"第{row_num}行: 客户名称不能超过200个字符")

    phone = row.get("联系电话", "").strip()
    if not phone:
        errors.append(f"第{row_num}行: 联系电话不能为空")
    elif not re.match(r'^[\d\-+\s()]{6,20}$', phone):
        errors.append(f"第{row_num}行: 联系电话格式不正确（请输入6-20位数字）")

    contact = row.get("联系人", "").strip()
    if len(contact) > 100:
        errors.append(f"第{row_num}行: 联系人不能超过100个字符")

    return errors


def parse_excel(
    file_bytes: bytes, entity_type: str
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """解析 Excel 文件，返回 (valid_rows, errors)"""
    tpl = TEMPLATES.get(entity_type)
    if not tpl:
        raise ValueError(f"不支持的实体类型: {entity_type}")

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], [{"row": 0, "message": "文件为空"}]

    headers = [str(h).strip() if h else "" for h in rows[0]]

    # 校验表头
    expected = tpl["headers"]
    if headers != expected:
        return [], [{
            "row": 0,
            "message": f"表头不匹配。预期: {' → '.join(expected)}，实际: {' → '.join(headers)}"
        }]

    valid_rows = []
    errors = []

    for i, row in enumerate(rows[1:], start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        row_dict = {}
        for col_idx, header in enumerate(headers):
            cell_val = row[col_idx] if col_idx < len(row) else None
            row_dict[header] = str(cell_val).strip() if cell_val is not None else ""

        # 执行校验
        if entity_type == "product":
            row_errors = _validate_product_row(row_dict, i)
        else:
            row_errors = _validate_customer_row(row_dict, i)

        if row_errors:
            errors.extend([{"row": i, "message": msg} for msg in row_errors])
        else:
            valid_rows.append(row_dict)

    return valid_rows, errors


def row_to_product_data(row: Dict[str, str]) -> Dict[str, Any]:
    price = None
    if row.get("默认单价", "").strip():
        try:
            price = Decimal(str(row["默认单价"]).strip())
        except InvalidOperation:
            price = None
    return {
        "name": row.get("产品名称", "").strip(),
        "model": row.get("规格型号", "").strip(),
        "default_price": price,
        "remark": row.get("备注", "").strip(),
    }


def row_to_customer_data(row: Dict[str, str]) -> Dict[str, Any]:
    return {
        "name": row.get("客户名称", "").strip(),
        "contact_person": row.get("联系人", "").strip(),
        "phone": row.get("联系电话", "").strip(),
        "address": row.get("地址", "").strip(),
        "remark": row.get("备注", "").strip(),
    }


# ---- 送货单 Excel 导出 ----
def export_delivery_note_xlsx(note) -> bytes:
    """纯代码生成送货单 Excel，不依赖外置模板"""
    from decimal import Decimal
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from backend.services.pdf_service import _amount_cn

    wb = Workbook()
    ws = wb.active
    ws.title = note.doc_number.replace("/", "_")

    # ---- 样式 ----
    cn_font = Font(name="Arial", size=12)
    title_font = Font(name="Arial", size=24, bold=True)
    header_font = Font(name="Arial", size=12, bold=True)
    bold_font = Font(name="Arial", size=12, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin_side = Side(style="thin", color="FF888888")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")

    # ---- 列宽 ----
    widths = {"A": 16.5, "B": 11.25, "C": 23.6, "D": 14.75, "E": 11.25, "F": 13.9, "G": 17.0, "H": 11.25}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for r in range(3, 19):
        ws.row_dimensions[r].height = 17.6

    # ---- 合并单元格 ----
    ws.merge_cells("A1:H2")  # 标题
    ws.merge_cells("B3:C3")  # 客户名称值
    ws.merge_cells("E3:F3")  # 开单日期值
    ws.merge_cells("B4:C4")  # 客户地址值
    ws.merge_cells("E4:F4")  # 单号值
    ws.merge_cells("A17:B17")  # 合计大写标签
    ws.merge_cells("C17:E17")  # 合计大写值
    ws.merge_cells("F17:G17")  # 合计小写标签

    # ---- 标题 ----
    ws["A1"] = "送货单"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    # ---- 基本信息行 ----
    info_rows = [
        (3, [("A", "客户名称:"), ("D", "开单日期："), ("G", "联系人:")]),
        (4, [("A", "客户地址:"), ("D", "单      号："), ("G", "电   话:")]),
    ]
    for row_num, cells in info_rows:
        for col, label in cells:
            c = ws[f"{col}{row_num}"]
            c.value = label
            c.font = cn_font
            c.alignment = center_align

    # 填充值
    ws["B3"] = note.customer_name or "";
    ws["B3"].font = cn_font;
    ws["B3"].alignment = center_align
    ws["E3"] = str(note.delivery_date);
    ws["E3"].font = cn_font;
    ws["E3"].alignment = center_align
    ws["H3"] = "";
    ws["H3"].font = cn_font;
    ws["H3"].alignment = center_align
    ws["B4"] = "";
    ws["B4"].font = cn_font;
    ws["B4"].alignment = center_align
    ws["E4"] = note.doc_number;
    ws["E4"].font = cn_font;
    ws["E4"].alignment = center_align
    ws["H4"] = "";
    ws["H4"].font = cn_font;
    ws["H4"].alignment = center_align

    # ---- 表头 ----
    headers = ["行号", "货品编码", "货品名称", "规格", "数量", "单价", "金额", "备注"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=ci, value=h)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    # ---- 明细行 ----
    total = Decimal("0")
    for idx, item in enumerate(note.items):
        row = 6 + idx
        if row > 16:
            break
        vals = [
            idx + 1,
            item.product_id,
            item.product_name or "",
            item.product_model or "",
            float(item.quantity),
            float(item.unit_price),
            float(item.subtotal),
            item.remark or "",
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = cn_font
            cell.border = thin_border
            if ci == 1:
                cell.alignment = center_align
            elif ci in (5, 6, 7):
                cell.alignment = right_align
                cell.number_format = "#,##0.00"
            else:
                cell.alignment = center_align
        total += item.subtotal

    # [以下为空] 标记 + 空行边框
    empty_start = 6 + len(note.items)
    if empty_start <= 16:
        # 后续空行加边框
        for r in range(empty_start, 17):
            for c in range(1, 9):
                if r == empty_start and c == 3:
                    cell = ws.cell(row=empty_start, column=c, value="以下为空")
                    cell.font = cn_font
                    cell.alignment = center_align
                ws.cell(row=r, column=c).border = thin_border
    else:
        for r in range(6 + len(note.items), 17):
            for c in range(1, 9):
                ws.cell(row=r, column=c).border = thin_border

    # ---- 合计行 ----
    ws["A17"] = "合计大写:"
    ws["A17"].font = bold_font;
    ws["A17"].alignment = center_align;
    ws["A17"].fill = header_fill;
    ws["A17"].border = thin_border
    ws["B17"].border = thin_border
    ws["C17"] = _amount_cn(total)
    ws["C17"].font = bold_font;
    ws["C17"].alignment = center_align;
    ws["C17"].fill = header_fill;
    ws["C17"].border = thin_border
    ws["D17"].border = thin_border
    ws["E17"].border = thin_border
    ws["F17"] = "合计小写:"
    ws["F17"].font = bold_font;
    ws["F17"].alignment = center_align;
    ws["F17"].fill = header_fill;
    ws["F17"].border = thin_border
    ws["G17"].border = thin_border
    ws["H17"] = float(total)
    ws["H17"].font = bold_font;
    ws["H17"].alignment = right_align;
    ws["H17"].fill = header_fill
    ws["H17"].border = thin_border;
    ws["H17"].number_format = "#,##0.00"

    # ---- 底部 ----
    ws["A18"] = "制单:"
    ws["A18"].font = cn_font;
    ws["A18"].alignment = center_align
    ws["G18"] = "负责人："
    ws["G18"].font = cn_font;
    ws["G18"].alignment = center_align

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()
    return buf.getvalue()



def generate_statement_xlsx(notes, start_date, end_date) -> bytes:
    """生成客户对账单 Excel。每客户一个 sheet，5列：日期 | 送货单号 | 明细 | 金额 | 备注"""
    from decimal import Decimal
    from collections import OrderedDict
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from backend.services.pdf_service import _amount_cn

    wb = Workbook()

    cn_font = Font(name="Arial", size=11)
    title_font = Font(name="Arial", size=20, bold=True)
    header_font = Font(name="Arial", size=11, bold=True)
    bold_font = Font(name="Arial", size=11, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_side = Side(style="thin", color="FF888888")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
    period_str = f"{start_date.isoformat()} 至 {end_date.isoformat()}"

    # Group notes by customer (preserve order)
    groups = OrderedDict()
    for note in notes:
        cid = note.customer_id
        if cid not in groups:
            groups[cid] = {"name": note.customer.name, "notes": []}
        groups[cid]["notes"].append(note)

    for idx, (cid, data) in enumerate(groups.items()):
        if idx == 0:
            ws = wb.active
            ws.title = data["name"][:31]
        else:
            ws = wb.create_sheet(title=data["name"][:31])

        # Column widths for 6 columns: A=序号, B=日期, C=送货单号, D=明细, E=金额, F=备注
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 16

        # Row 1: Title (merged A1:E1)
        ws.merge_cells("A1:F1")
        ws["A1"] = "客户对账单"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align

        # Row 2: Customer name & period (merged A2:E2)
        ws.merge_cells("A2:F2")
        ws["A2"] = f"客户：{data['name']}    对账期间：{period_str}"
        ws["A2"].font = Font(name="Arial", size=12, bold=True)
        ws["A2"].alignment = center_align

        # Row 3: Optional customer info (address / contact / phone)
        if notes:
            info_parts = []
            for field, label in [("address", "地址"), ("contact_person", "联系人"), ("phone", "电话")]:
                val = getattr(notes[0].customer, field, None)
                if val:
                    info_parts.append(f"{label}：{val}")
            if info_parts:
                ws.merge_cells("A3:F3")
                ws["A3"] = "  ".join(info_parts)
                ws["A3"].font = cn_font
                ws["A3"].alignment = left_align
                data_start_row = 4
            else:
                data_start_row = 3
        else:
            data_start_row = 3

        # Header row (6 columns: 序号, 日期, 送货单号, 明细, 金额, 备注)
        headers = ["序号", "日期", "送货单号", "明细", "金额", "备注"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=data_start_row, column=ci, value=h)
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = thin_border

        # Data rows — one note per group, items expand into sub-rows
        grand_total = Decimal("0")
        row = data_start_row  # current row tracker
        for ri, note in enumerate(data["notes"]):
            # Compute item rows needed
            items = [item for item in note.items if item.product]
            # Group header row (date, doc_number, total_amount, remark)
            row += 1
            date_str = note.delivery_date.isoformat() if hasattr(note.delivery_date, "isoformat") else str(note.delivery_date)
            note_amount = float(note.total_amount or 0)
            # 序号 | 日期 | 送货单号 | (空) | 总金额 | 备注
            header_vals = [ri + 1, date_str, note.doc_number, "", note_amount, note.remark or ""]
            for ci, v in enumerate(header_vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.font = cn_font
                cell.border = thin_border
                if ci == 5:  # 金额
                    cell.alignment = right_align
                    cell.number_format = '#,##0.00'
                else:
                    cell.alignment = center_align
            ws.row_dimensions[row].height = 20
            # Item sub-rows
            for item in items:
                row += 1
                subtotal = float(item.subtotal) if hasattr(item, 'subtotal') and item.subtotal else 0
                # Clean number formatting
                def _trim(s):
                    if "." in s:
                        return s.rstrip("0").rstrip(".")
                    return s
                price_s = _trim(str(item.unit_price))
                qty_s = _trim(str(item.quantity))
                detail_str = f"{item.product.name} * {price_s} * {qty_s}"
                # (空) | (空) | (空) | 明细 | 小计 | (空)
                item_vals = ["", "", "", detail_str, subtotal, ""]
                for ci, v in enumerate(item_vals, 1):
                    cell = ws.cell(row=row, column=ci, value=v)
                    cell.font = cn_font
                    cell.border = thin_border
                    if ci == 5:  # 金额
                        cell.alignment = right_align
                        cell.number_format = '#,##0.00'
                    elif ci == 4:  # 明细 — left align
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align
                ws.row_dimensions[row].height = 20
            grand_total += note.total_amount or Decimal("0")

        # Total row — 大写 + 小写
        total_row = row + 1
        ws.merge_cells(f"A{total_row}:D{total_row}")
        total_cn = _amount_cn(grand_total)
        total_label_cell = ws.cell(row=total_row, column=1, value=f"合计金额（大写）：{total_cn}")
        total_label_cell.font = bold_font
        total_label_cell.alignment = left_align
        total_label_cell.fill = header_fill
        total_label_cell.border = thin_border
        ws.cell(row=total_row, column=2).border = thin_border
        ws.cell(row=total_row, column=2).fill = header_fill
        ws.cell(row=total_row, column=3).border = thin_border
        ws.cell(row=total_row, column=3).fill = header_fill
        ws.cell(row=total_row, column=4).border = thin_border
        ws.cell(row=total_row, column=4).fill = header_fill

        total_label_cell2 = ws.cell(row=total_row, column=5, value="小写：")
        total_label_cell2.font = bold_font
        total_label_cell2.alignment = Alignment(horizontal="right", vertical="center")
        total_label_cell2.fill = header_fill
        total_label_cell2.border = thin_border
        total_val_cell = ws.cell(row=total_row, column=6, value=float(grand_total))
        total_val_cell.font = bold_font
        total_val_cell.alignment = right_align
        total_val_cell.fill = header_fill
        total_val_cell.number_format = '#,##0.00'
        total_val_cell.border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()
    return buf.getvalue()
