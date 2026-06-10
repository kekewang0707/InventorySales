"""对账单功能测试"""
import sys
from pathlib import Path

project_root = Path(__file__).resolve().parent.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from datetime import date
from decimal import Decimal

from backend.schemas.statement import StatementQuery
from backend.services.excel_service import generate_statement_xlsx


class FakeProduct:
    def __init__(self, name):
        self.name = name


class FakeItem:
    def __init__(self, product_name, quantity, unit_price=None, subtotal=None):
        self.product = FakeProduct(product_name)
        self.quantity = quantity
        self.unit_price = unit_price or quantity
        self.subtotal = subtotal or (quantity * (unit_price or quantity))


class FakeCustomer:
    def __init__(self, name, address="", contact_person="", phone=""):
        self.name = name
        self.address = address
        self.contact_person = contact_person
        self.phone = phone


class FakeNote:
    def __init__(self, customer_id, customer_name, doc_number, delivery_date, total_amount, remark="", items=None):
        self.customer_id = customer_id
        self.customer = FakeCustomer(customer_name)
        self.doc_number = doc_number
        self.delivery_date = delivery_date
        self.total_amount = total_amount
        self.remark = remark
        self.items = items or []


# ---- 测试 _build_statements 逻辑 ----

def test_query_statements_single_customer():
    """验证单个客户对账单查询，正确合并到一组。"""
    from backend.routers.statements import _build_statements

    notes = [
        FakeNote(1, "客户A", "DN001", date(2026, 6, 1), Decimal("100.00"), "", [
            FakeItem("产品X", Decimal("2"))
        ]),
        FakeNote(1, "客户A", "DN002", date(2026, 6, 5), Decimal("200.00"), "备注", [
            FakeItem("产品Y", Decimal("1"))
        ]),
    ]

    statements = _build_statements(notes)
    assert len(statements) == 1
    assert statements[0].customer_name == "客户A"
    assert len(statements[0].delivery_notes) == 2
    assert statements[0].total_amount == Decimal("300.00")


def test_query_statements_multiple_customers():
    """验证多客户对账单查询，正确按客户分组。"""
    from backend.routers.statements import _build_statements

    notes = [
        FakeNote(1, "客户A", "DN001", date(2026, 6, 1), Decimal("100.00")),
        FakeNote(2, "客户B", "DN002", date(2026, 6, 2), Decimal("200.00")),
        FakeNote(1, "客户A", "DN003", date(2026, 6, 3), Decimal("300.00")),
    ]

    statements = _build_statements(notes)
    assert len(statements) == 2

    stmt_a = [s for s in statements if s.customer_name == "客户A"][0]
    assert stmt_a.total_amount == Decimal("400.00")
    assert len(stmt_a.delivery_notes) == 2

    stmt_b = [s for s in statements if s.customer_name == "客户B"][0]
    assert stmt_b.total_amount == Decimal("200.00")
    assert len(stmt_b.delivery_notes) == 1


def test_query_statements_empty():
    """验证无数据时返回空列表。"""
    from backend.routers.statements import _build_statements
    statements = _build_statements([])
    assert statements == []


# ---- 测试 Excel 生成 ----

def test_generate_statement_xlsx_single():
    """验证单个客户对账单 Excel 生成。"""
    notes = [
        FakeNote(1, "客户A", "DN001", date(2026, 6, 1), Decimal("150.00"), "正常", [
            FakeItem("产品X", Decimal("3"))
        ]),
    ]

    xlsx_bytes = generate_statement_xlsx(notes, date(2026, 6, 1), date(2026, 6, 30))
    assert xlsx_bytes is not None
    assert len(xlsx_bytes) > 0

    # Verify it's a valid xlsx
    from openpyxl import load_workbook
    import io
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.active.title == "客户A"
    assert wb.active["A1"].value == "客户对账单"
    wb.close()


def test_generate_statement_xlsx_multiple():
    """验证多个客户对账单 Excel 生成（每人一个 sheet）。"""
    notes = [
        FakeNote(1, "客户A", "DN001", date(2026, 6, 1), Decimal("100.00")),
        FakeNote(2, "客户B", "DN002", date(2026, 6, 2), Decimal("200.00")),
    ]

    xlsx_bytes = generate_statement_xlsx(notes, date(2026, 6, 1), date(2026, 6, 30))
    assert xlsx_bytes is not None

    from openpyxl import load_workbook
    import io
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    assert len(wb.sheetnames) == 2
    assert "客户A" in wb.sheetnames
    assert "客户B" in wb.sheetnames
    wb.close()


# ---- 测试 StatementQuery schema ----

def test_statement_query_defaults():
    """验证 StatementQuery 默认值行为。"""
    q = StatementQuery(start_date=date(2026, 6, 1), end_date=date(2026, 6, 30))
    assert q.customer_id is None
    assert q.start_date == date(2026, 6, 1)
    assert q.end_date == date(2026, 6, 30)


def test_statement_query_with_customer():
    """验证指定客户 ID 的查询参数。"""
    q = StatementQuery(customer_id=5, start_date=date(2026, 6, 1), end_date=date(2026, 6, 30))
    assert q.customer_id == 5
